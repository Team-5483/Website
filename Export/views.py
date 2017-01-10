from django.shortcuts import render
from openpyxl import load_workbook
from Inventory.models import Item
from django.shortcuts import HttpResponseRedirect


# Create your views here.


def index(request):
    file_not_found = ''
    if request.method == 'POST':
        file_path = request.POST.get('file_path', None)
        if file_path is not None and file_path is not '':
            try:
                export_excel(file_path)
                return HttpResponseRedirect('/Inventory/')
            except FileNotFoundError:
                file_not_found = 'File Not Found'

    return render(request, "Export/index.html", {'file_not_found': file_not_found})

def export_excel(fileName):
    wb = load_workbook(fileName)
    ws = wb.active

    for row in ws.iter_rows(min_row=1, max_col=8, max_row=len(Item.objects)):
        for cell in row:
            cell.append(Item.objects.get_all)



