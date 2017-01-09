from django.shortcuts import render
from openpyxl import load_workbook
from Inventory.models import Item
from django.shortcuts import HttpResponseRedirect


# Create your views here.


def index(request):
    file_not_found = ''
    if request.method == 'POST':
        file_path = request.POST.get('file_path', None)
        if file_path is not None and file_path is not '':
            try:
                excel = import_excel(file_path)
                save(excel[0], excel[1])
                return HttpResponseRedirect('/Inventory/')
            except FileNotFoundError:
                file_not_found = 'File Not Found'

    return render(request, "ExcelParser/index.html", {'file_not_found' : file_not_found})


def import_excel(fileName):
    wb = load_workbook(fileName)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=1, max_col=8, max_row=ws.max_row):
        for cell in row:
            items.append(cell.value)

    return items, ws


def data_entry(nameParam="", categoryParam="", sourceParam="", measurmentParam="", referenceParam="",
               ppuParam=0, priceParam=0, quantityParam=0):
    if not nameParam.startswith("#"):
        new_item = Item()
        new_item.item_name = nameParam
        new_item.item_category = categoryParam
        new_item.item_source = sourceParam
        new_item.item_measurement = measurmentParam
        new_item.item_reference = referenceParam
        new_item.item_unit_price = ppuParam
        new_item.item_total_price = priceParam
        new_item.item_quantity = quantityParam
        new_item.save()


def save(items, ws):
    for i in range(ws.max_row):
        data_entry(items[(i * 8)], items[(i * 8) + 1], items[(i * 8) + 2], items[(i * 8) + 3], items[(i * 8) + 4],
                                    items[(i * 8) + 5], items[(i * 8) + 6], items[(i * 8) + 7])
