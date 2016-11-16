from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import sqlite3
from openpyxl.utils import get_column_letter

conn = sqlite3.connect('test.db')
c = conn.cursor()


def import_excel(excelName=""):

    wb = load_workbook(excelName)
    ws = wb.active
    i = 0

    nameInCell=""
    categoryInCell=""
    sourceInCell=""
    measurmentInCell=""
    referenceInCell=""
    ppuInCell=""
    priceInCell=""
    quantityInCell=""

    items = []

    def create_table():
         c.execute(
             'CREATE TABLE IF NOT EXISTS items(name TEXT, category TEXT, source TEXT, measurment TEXT, reference TEXT, ppu INT, price INT, quantity INT)')

    def data_entry(nameParam="", categoryParam="", sourceParam="", measurmentParam="", referenceParam="",
                      ppuParam=0, priceParam=0, quantityParam=0):
        c.execute(
            "INSERT INTO items (name, category, source, measurment, reference, ppu, price, quantity) VALUES(?, ?, ?, ?, ?, ?, ?, ?)",
                (nameParam, categoryParam, sourceParam, measurmentParam, referenceParam, ppuParam, priceParam,
                 quantityParam))
        conn.commit()

    create_table()

    for row in ws.iter_rows(min_row=1, max_col=8, max_row=10):
        for cell in row:
            items.append(cell.value)

    for i in range(10):
        data_entry(items[(i*8)], items[(i*8)+1], items[(i*8)+2], items[(i*8)+3], items[(i*8)+4], items[(i*8)+5], items[(i*8)+6], items[(i*8)+7])

import_excel('test.xlsx')

c.close()
conn.close()